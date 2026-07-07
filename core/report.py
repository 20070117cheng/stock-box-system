# -*- coding: utf-8 -*-
"""Excel 回測報告產生器（版面沿用舊腳本《批次回測報告.py》）。"""
import io
import math

import matplotlib

matplotlib.use("Agg")  # 無視窗環境（Actions / Streamlit）必須用 Agg
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.backtest import BacktestResult

plt.rcParams["font.sans-serif"] = ["Microsoft JhengHei", "Arial Unicode MS",
                                   "Noto Sans CJK TC", "SimHei"]
plt.rcParams["axes.unicode_minus"] = False

HEADERS = ["日期", "日收盤價", "K值", "D值", "平均持股成本", "移動停利線",
           "固定停損線", "正宗箱頂(高)", "正宗箱底(低)", "持股張數",
           "預計進場價", "預計出場價", "訊號狀態", "進出場原因說明", "損益金額"]


def _chart_png(daily, stock_id: str) -> io.BytesIO:
    fig, ax = plt.subplots(figsize=(10, 5.5))
    x = range(len(daily))

    ax.plot(x, daily["close"], label="日收盤價", color="black", linewidth=1.5)
    ax.plot(x, daily["移動停利線"], label="移動停利線", color="green", linestyle="--")
    ax.plot(x, daily["固定停損線"], label="固定停損線", color="blue", linestyle="--")
    ax.plot(x, daily["前箱高"], label="正宗箱頂", color="red", linewidth=1.0)
    ax.plot(x, daily["前箱低"], label="正宗箱底", color="red", linewidth=1.0)
    ax.fill_between(x, daily["前箱高"], daily["前箱低"], step="post",
                    color="red", alpha=0.03)

    last_idx = len(daily) - 1
    ax.text(last_idx, daily["close"].iloc[-1],
            f" 現價: {daily['close'].iloc[-1]:.2f}", color="black")
    ax.text(last_idx, daily["前箱高"].iloc[-1],
            f" 箱頂: {daily['前箱高'].iloc[-1]:.2f}", color="red")
    ax.text(last_idx, daily["前箱低"].iloc[-1],
            f" 箱底: {daily['前箱低'].iloc[-1]:.2f}", color="red")

    max_labels = 6
    step = max(1, math.ceil(len(daily) / max_labels))
    ticks = list(range(0, len(daily), step))
    if last_idx not in ticks:
        ticks.append(last_idx)
    ax.set_xticks(ticks)
    ax.set_xticklabels([daily.index[i].strftime("%Y-%m-%d") for i in ticks],
                       rotation=0, ha="center", fontsize=9)

    ax.set_title(f"{stock_id} 正宗達瓦斯箱型理論與策略回測圖",
                 fontsize=12, pad=8, fontweight="bold")
    ax.grid(True, alpha=0.3)
    ax.legend(loc="upper left", fontsize=8)
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, bbox_inches="tight", dpi=120, format="png")
    buf.seek(0)
    plt.close(fig)
    return buf


def build_excel(result: BacktestResult, stock_id: str, stock_name: str) -> bytes:
    """回傳 xlsx 檔內容 bytes（供下載或存檔）。"""
    daily = result.daily
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "策略回測報告"

    font_header = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Microsoft JhengHei", size=10)
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_alert = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_red = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    border_thin = Border(left=Side(style="thin", color="D9D9D9"),
                         right=Side(style="thin", color="D9D9D9"),
                         top=Side(style="thin", color="D9D9D9"),
                         bottom=Side(style="thin", color="D9D9D9"))

    ws.merge_cells("A1:O1")
    ws["A1"] = f"【明日交易提示】 {result.tomorrow_desc}"
    ws["A1"].font = Font(name="Microsoft JhengHei", size=11, bold=True, color="C00000")
    ws["A1"].fill = fill_alert
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35

    for col_num, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    ws.row_dimensions[3].height = 45

    for row_num, (idx, row) in enumerate(daily.iterrows(), 4):
        ws.cell(row=row_num, column=1, value=idx.strftime("%Y-%m-%d")).alignment = align_center
        ws.cell(row=row_num, column=2, value=round(row["close"], 2)).alignment = align_right
        ws.cell(row=row_num, column=3, value=round(row["k"], 2)).alignment = align_right
        ws.cell(row=row_num, column=4, value=round(row["d"], 2)).alignment = align_right
        ws.cell(row=row_num, column=5, value=round(row["平均成本"], 2)).alignment = align_right
        ws.cell(row=row_num, column=6, value=round(row["移動停利線"], 2)).alignment = align_right
        ws.cell(row=row_num, column=7, value=round(row["固定停損線"], 2)).alignment = align_right
        ws.cell(row=row_num, column=8, value=round(row["前箱高"], 2)).alignment = align_right
        ws.cell(row=row_num, column=9, value=round(row["前箱低"], 2)).alignment = align_right
        ws.cell(row=row_num, column=10, value=int(row["持股張數"])).alignment = align_center
        ws.cell(row=row_num, column=11,
                value=round(row["進場目標價"], 2) if row["進場目標價"] > 0 else "-"
                ).alignment = align_right
        ws.cell(row=row_num, column=12,
                value=round(row["出場目標價"], 2) if row["出場目標價"] > 0 else "-"
                ).alignment = align_right

        status_text = row["訊號狀態"]
        status_cell = ws.cell(row=row_num, column=13, value=status_text)
        status_cell.alignment = align_center
        reason_cell = ws.cell(row=row_num, column=14, value=row["進出場原因說明"])
        reason_cell.alignment = align_left

        if status_text in ("買入進場", "重返進場", "加碼"):
            status_cell.fill = fill_blue
            reason_cell.fill = fill_blue
        elif status_text == "出場":
            status_cell.fill = fill_red
            reason_cell.fill = fill_red

        amt_cell = ws.cell(row=row_num, column=15, value=round(row["損益金額"], 0))
        amt_cell.alignment = align_right
        if row["損益金額"] > 0:
            amt_cell.font = Font(name="Microsoft JhengHei", size=10,
                                 color="FF0000", bold=True)
        elif row["損益金額"] < 0:
            amt_cell.font = Font(name="Microsoft JhengHei", size=10, color="008000")

        for col_num in range(1, 16):
            c = ws.cell(row=row_num, column=col_num)
            c.font = c.font if col_num == 15 else font_data
            c.border = border_thin
        ws.row_dimensions[row_num].height = 20

    xl_img = Image(_chart_png(daily, stock_id))
    xl_img.width = 660
    xl_img.height = 300
    ws.add_image(xl_img, "Q2")

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter == "A":
            ws.column_dimensions[col_letter].width = 14
        elif col_letter in list("BCDEFGHIJKL"):
            ws.column_dimensions[col_letter].width = 7
        elif col_letter in ("M", "O"):
            ws.column_dimensions[col_letter].width = 11
        elif col_letter == "N":
            ws.column_dimensions[col_letter].width = 48

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            current = cell.alignment or Alignment()
            cell.alignment = Alignment(wrap_text=True,
                                       horizontal=current.horizontal,
                                       vertical=current.vertical)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
