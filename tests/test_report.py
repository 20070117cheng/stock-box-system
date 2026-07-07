# -*- coding: utf-8 -*-
import io

import numpy as np
import openpyxl
import pandas as pd

from core.backtest import run_backtest
from core.config import DEFAULTS
from core.report import HEADERS, build_excel


def _small_result():
    i = np.arange(300)
    closes = 100 + 0.35 * i + 6 * np.sin(i / 5)
    dates = pd.bdate_range("2025-01-02", periods=len(closes))
    df = pd.DataFrame(
        {"open": closes, "high": closes * 1.01, "low": closes * 0.99,
         "close": closes, "volume": [5000] * len(closes)}, index=dates)
    start = df.index[260].strftime("%Y-%m-%d")
    return run_backtest(df, start, dict(DEFAULTS))


def test_build_excel_layout():
    res = _small_result()
    data = build_excel(res, "2330", "台積電")
    assert isinstance(data, bytes) and len(data) > 0

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.title == "策略回測報告"
    assert "明日交易提示" in str(ws["A1"].value)
    headers = [ws.cell(row=3, column=c).value for c in range(1, len(HEADERS) + 1)]
    assert headers == HEADERS
    # 資料列數 = 回測天數
    assert ws.cell(row=4, column=1).value is not None
    n_rows = 0
    r = 4
    while ws.cell(row=r, column=1).value is not None:
        n_rows += 1
        r += 1
    assert n_rows == len(res.daily)
